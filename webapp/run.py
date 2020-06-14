from flask import Flask, render_template
import folium
import os
import json
import time
from app_functions import dijkstra, get_coordinates, dijkstra_endlist, shortestpath, searchbasedRS, grab_share, get_largest_angle, load_object
from openpyxl import Workbook
from openpyxl import load_workbook

app = Flask(__name__)

@app.route('/')
def index():
    # UP Diliman coordinates
    start_coords = (14.6538, 121.0685)

    # Geojson data
    route = os.path.join('map_data', 'route.json')

    # initialize excel file for data writing
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Case number'
    ws['B1'] = '# of Passengers'
    ws['C1'] = 'Radial Parameter Pass (m)'
    ws['D1'] = '# of Drivers'
    ws['E1'] = 'Radial Parameter Driver (m)'
    ws['F1'] = 'Route Length (km)'
    ws['G1'] = 'Route Time (minutes)'
    ws['H1'] = '# of Matches'
    ws['I1'] = 'Trip Price (Php)'
    ws['J1'] = 'SRPs'
    ws['K1'] = 'SRPs'
    ws['L1'] = 'SRPs'
    ws['M1'] = 'Average SRP'
    ws['N1'] = 'Processing time (seconds)'
    ws['O1'] = 'Route Length (km)'
    ws['P1'] = 'Route Time (minutes)'
    ws['Q1'] = '# of Matches'
    ws['R1'] = 'Trip Price (Php)'
    ws['S1'] = 'SRPs'
    ws['T1'] = 'SRPs'
    ws['U1'] = 'SRPs'
    ws['V1'] = 'Average SRP'
    ws['W1'] = 'Angle'
    ws['X1'] = 'Processing time (seconds)'

    # passengers = ['30763177','5585622052','30763115']
    # passenger_destinations = ['17216442','2517360527','2517360522']
    # drivers = ['30763220','1402297896','5499240548']
    # passengers = ['5446811375','29025439','17216409']
    # passenger_destinations = ['17216442','2517360527','2517360522']
    # drivers = ['30763220','1402297896','5499240548']

    passengers = load_object('passenger_sources.pkl')
    passenger_destinations = load_object('passenger_destinations.pkl')
    drivers = load_object('driver_locations.pkl')
    pass_radii = load_object('rand_pass_radii.pkl')
    dri_radii = load_object('rand_dri_radii.pkl')

    # ctr = 0
    row_excel = 2
    main_ctr = 0
    while main_ctr < 300:
        a = ws.cell(row = row_excel, column = 1)
        a.value = main_ctr+1
        a = ws.cell(row = row_excel, column = 2)
        a.value = len(passengers[main_ctr])
        a = ws.cell(row = row_excel, column = 3)
        a.value = pass_radii[main_ctr]
        a = ws.cell(row = row_excel, column = 4)
        a.value = len(drivers[main_ctr])
        a = ws.cell(row = row_excel, column = 5)
        a.value = dri_radii[main_ctr]

        folium_map = folium.Map(location=start_coords,zoom_start=16,height='85%')
        # searchbased-rs implementation ------------------
        start_time = time.perf_counter()
        sources, destinations, path, route_distance, route_time, fare, srp_list = searchbasedRS('adj_list_obj.pkl', drivers[main_ctr], passengers[main_ctr], passenger_destinations[main_ctr], 0.5, row_excel)
        end_time = time.perf_counter()
        print('\nSearch-BasedRS time elapsed (seconds): ' + str(end_time - start_time) + '\n')
        if sources != None:
            a = ws.cell(row = row_excel, column = 6)
            a.value = route_distance
            a = ws.cell(row = row_excel, column = 7)
            a.value = route_time
            a = ws.cell(row = row_excel, column = 8)
            a.value = len(sources)
            a = ws.cell(row = row_excel, column = 9)
            a.value = fare
            srp_column = 10
            for srp in srp_list:
                a = ws.cell(row = row_excel, column = srp_column)
                a.value = srp
                srp_column += 1
            a = ws.cell(row = row_excel, column = 14)
            a.value = end_time - start_time
        # ------------------------------------------------

        # save html file for searchbased
        # ----------------------------------------------------------------------- #
            driver_coordinates = get_coordinates('nodes_coordinates.pkl', drivers[main_ctr])
            for driver_coordinate in driver_coordinates:
                    # folium.Marker(driver_coordinate, tooltip='Driver', icon=folium.Icon(color='red', icon='user')).add_to(folium_map),
                     # Create custom marker icon
                    car_icon = folium.features.CustomIcon('map_data/car_marker.png', icon_size=(40, 40))
                    folium.Marker(driver_coordinate,icon=car_icon).add_to(folium_map),

            # # add additional markers and path if match/es found
            if sources!=None:
                # add path
                coordinates = get_coordinates('nodes_coordinates.pkl', path)
                way_sample=folium.PolyLine(locations=coordinates,weight=5,color = 'red')
                folium_map.add_child(way_sample)
                # Create markers for matched sources and destinations
                # markers for other available passengers shown
                # All drivers will have a marker to mimic actual ride sharing applications
                coordinates_pass = get_coordinates('nodes_coordinates.pkl', passengers[main_ctr])
                for passenger in coordinates_pass:
                    if passenger in coordinates:
                        continue
                    icon_color = 'beige'
                    folium.Marker(passenger, tooltip='Source', icon=folium.Icon(color=icon_color, icon='user')).add_to(folium_map),

                ctr = 0
                for source in sources:
                    index = path.index(source)
                    if ctr == 0:
                        icon_color = 'blue'
                    elif ctr == 1:
                        icon_color = 'green'
                    elif ctr == 2:
                        icon_color = 'purple'
                    folium.Marker(coordinates[index], tooltip='Source', icon=folium.Icon(color=icon_color, icon='chevron-up')).add_to(folium_map),
                    ctr+=1

                ctr = 0
                for destination in destinations:
                    index = path.index(destination)
                    if ctr == 0:
                        icon_color = 'blue'
                    elif ctr == 1:
                        icon_color = 'green'
                    elif ctr == 2:
                        icon_color = 'purple'
                    folium.Marker(coordinates[index], tooltip='Destination', icon=folium.Icon(color=icon_color, icon='chevron-down')).add_to(folium_map),
                    ctr+=1

        # # Geojson overlay
        # folium.GeoJson(route, name='route').add_to(folium_map)

        fn='templates/map.html'
        fn = fn.replace(fn, 'templates/map' + str(main_ctr) + '_search.html')
        folium_map.save(fn)

        folium_map = folium.Map(location=start_coords,zoom_start=16,height='85%')
        # # grab algorithm implementation ------------------
        start_time = time.perf_counter()
        sources, destinations, path, route_distance, route_time, fare, srp_list, angle = grab_share('adj_list_obj.pkl', drivers[main_ctr], passengers[main_ctr], passenger_destinations[main_ctr], 60)
        end_time = time.perf_counter()
        print('\nGrabShare Algorithm time elapsed (seconds): ' + str(end_time - start_time) + '\n')
        if sources != None:
            a = ws.cell(row = row_excel, column = 15)
            a.value = route_distance
            a = ws.cell(row = row_excel, column = 16)
            a.value = route_time
            a = ws.cell(row = row_excel, column = 17)
            a.value = len(sources)
            a = ws.cell(row = row_excel, column = 18)
            a.value = fare
            srp_column = 19
            for srp in srp_list:
                a = ws.cell(row = row_excel, column = srp_column)
                a.value = srp
                srp_column += 1
            a = ws.cell(row = row_excel, column = 23)
            a.value = angle
            a = ws.cell(row = row_excel, column = 24)
            a.value = end_time - start_time

            # save html file for grab
            # ----------------------------------------------------------------------- #
            driver_coordinates = get_coordinates('nodes_coordinates.pkl', drivers[main_ctr])
            for driver_coordinate in driver_coordinates:
                    # folium.Marker(driver_coordinate, tooltip='Driver', icon=folium.Icon(color='red', icon='user')).add_to(folium_map),
                     # Create custom marker icon
                    car_icon = folium.features.CustomIcon('map_data/car_marker.png', icon_size=(40, 40))
                    folium.Marker(driver_coordinate,icon=car_icon).add_to(folium_map),

            # # add additional markers and path if match/es found
            if sources!=None:
                # add path
                coordinates = get_coordinates('nodes_coordinates.pkl', path)
                way_sample=folium.PolyLine(locations=coordinates,weight=5,color = 'red')
                folium_map.add_child(way_sample)
                # Create markers for matched sources and destinations
                # markers for other available passengers shown
                # All drivers will have a marker to mimic actual ride sharing applications
                coordinates_pass = get_coordinates('nodes_coordinates.pkl', passengers[main_ctr])
                for passenger in coordinates_pass:
                    if passenger in coordinates:
                        continue
                    icon_color = 'beige'
                    folium.Marker(passenger, tooltip='Source', icon=folium.Icon(color=icon_color, icon='user')).add_to(folium_map),

                ctr = 0
                for source in sources:
                    index = path.index(source)
                    if ctr == 0:
                        icon_color = 'blue'
                    elif ctr == 1:
                        icon_color = 'green'
                    elif ctr == 2:
                        icon_color = 'purple'
                    folium.Marker(coordinates[index], tooltip='Source', icon=folium.Icon(color=icon_color, icon='chevron-up')).add_to(folium_map),
                    ctr+=1

                ctr = 0
                for destination in destinations:
                    index = path.index(destination)
                    if ctr == 0:
                        icon_color = 'blue'
                    elif ctr == 1:
                        icon_color = 'green'
                    elif ctr == 2:
                        icon_color = 'purple'
                    folium.Marker(coordinates[index], tooltip='Destination', icon=folium.Icon(color=icon_color, icon='chevron-down')).add_to(folium_map),
                    ctr+=1

        # # Geojson overlay
        # folium.GeoJson(route, name='route').add_to(folium_map)

        fn='templates/map.html'
        fn = fn.replace(fn, 'templates/map' + str(main_ctr) + '_grab.html')
        folium_map.save(fn)

        # # ------------------------------------------------
        row_excel += 1
        main_ctr += 1


    wb.save('data.xlsx')
    return render_template('index.html')

if __name__ == '__main__':
    app.run(debug=True)

